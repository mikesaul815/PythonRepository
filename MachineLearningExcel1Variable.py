from openpyxl import load_workbook
from sklearn.ensemble import RandomForestRegressor
import numpy as np

# Step 1: Read the Excel file
excel_file = r'filepath\TestData.xlsx'  #Update File path with where you are keeping TestData.xlsx
                                        #Update the file name TestData.xlsx to whatever file you are working on
workbook = load_workbook(excel_file)
sheet = workbook.active

# Step 2: Extract data from the Excel sheet
data = list(sheet.iter_rows(values_only=True))

# Extract variable names
variable_names = data[0]
outcome_index = variable_names.index('Outcome')
column1_index = variable_names.index('Column1')

# Prepare data for machine learning
X_train = []
y_train = []
for row in data[1:]:
    try:
        column1_value = float(row[column1_index])
        outcome_value = float(row[outcome_index])
    except ValueError:
        continue  # Skip rows with missing or non-numeric values

    X_train.append([column1_value])
    y_train.append(outcome_value)

X_train = np.array(X_train)
y_train = np.array(y_train)

# Step 3: Train your machine learning model (Random Forest Regressor) using 'Column1' to predict 'Outcome'
model = RandomForestRegressor()
model.fit(X_train, y_train)

# Step 4: Make predictions for all rows
predictions = model.predict(X_train)

# Step 5: Write the predicted outcome and 'Column1' values to the Excel file
start_column = 7  # G corresponds to column number 7 in Excel
differences = []
for i, value in enumerate(predictions):
    row_index = i + 1  # Index starts from 1 in Excel
    predicted_value = value
    actual_value = data[row_index][outcome_index]
    difference = actual_value - predicted_value
    squared_difference = difference ** 2
    differences.append(squared_difference)
    
    sheet.cell(row=row_index + 1, column=start_column, value=predicted_value)  # Writing predicted outcome
    sheet.cell(row=row_index + 1, column=start_column + 1, value=data[row_index][column1_index])  # Writing 'Column1' value
    sheet.cell(row=row_index + 1, column=start_column + 2, value=difference)  # Writing difference in column I
    sheet.cell(row=row_index + 1, column=start_column + 3, value=squared_difference)  # Writing squared difference in column J

# Add labels to the new columns
sheet.cell(row=1, column=start_column, value="PredictedOutcome")  # Add label above PredictedOutcome
sheet.cell(row=1, column=start_column + 1, value="Column1")  # Add label above 'Column1'
sheet.cell(row=1, column=start_column + 2, value="Difference")  # Add label above 'Difference'
sheet.cell(row=1, column=start_column + 3, value="SquaredDifference")  # Add label above 'SquaredDifference'

# Step 6: Calculate and write the RMSE value
rmse = np.sqrt(np.mean(differences))
sheet.cell(row=1, column=13, value="RMSE")  # Writing 'RMSE' in cell M1 (column 13)
sheet.cell(row=1, column=14, value=rmse)    # Writing RMSE value in cell N1 (column 14)

# Step 7: Save the updated Excel file
workbook.save(excel_file)
